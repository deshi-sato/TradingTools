"""
db_updater.py
Auto-specs CI test: added module docstring. (v2)
仕様固定版 + SNAPSHOT_MAP 対応
ダミー行: 自動生成テスト用
"""
from __future__ import annotations
import re
import sqlite3
import time
from datetime import datetime, timedelta, time as dtime
from pathlib import Path
from typing import Dict, Tuple
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Tuple, Callable, Any

# ========= 設定 =========
SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = SCRIPT_DIR / "デイトレ株価データ.xlsm"
DB_PATH = SCRIPT_DIR / "data" / "rss_data.db"

# ========= DB =========
DDL = """
-- 1分足
CREATE TABLE IF NOT EXISTS minute_data (
    ticker      TEXT    NOT NULL,   -- 銘柄コード(16進4桁・大文字)
    sheet_name  TEXT    NOT NULL,   -- 銘柄名称（シート名）
    datetime    TEXT    NOT NULL,   -- 'YYYY-MM-DD HH:MM:SS' (JST)
    open        REAL,
    high        REAL,
    low         REAL,
    close       REAL,
    volume      INTEGER,
    PRIMARY KEY (ticker, datetime)
);
CREATE INDEX IF NOT EXISTS idx_minute_data_ticker_datetime
  ON minute_data (ticker, datetime);

-- 最新スナップショット（銘柄ごとに1行）
CREATE TABLE IF NOT EXISTS quote_latest (
  ticker       TEXT    NOT NULL,      -- 16進4桁（大文字）
  sheet_name   TEXT    NOT NULL,      -- シート名（銘柄名）
  last         REAL,                  -- 現在値
  prev_close   REAL,                  -- 前日終値
  open         REAL,
  high         REAL,
  low          REAL,
  volume       INTEGER,               -- 出来高（累計）
  turnover     INTEGER,               -- 売買代金（累計）
  diff         REAL,                  -- 前日比
  diff_pct     REAL,                  -- 前日比率[%]
  updated_at   TEXT    NOT NULL,      -- 'YYYY-MM-DD HH:MM:SS' (JST)
  PRIMARY KEY (ticker)
);
CREATE INDEX IF NOT EXISTS idx_quote_latest_updated_at
  ON quote_latest (updated_at);
"""


# ========= 共通ユーティリティ =========
def _to_float(x):
    """数値/数値文字列→float（それ以外は None）"""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "")
    return float(s) if re.fullmatch(r"-?\d+(\.\d+)?", s) else None


def _to_int(x):
    """数値/数値文字列→int（それ以外は None）"""
    f = _to_float(x)
    return None if f is None else int(f)


def _timestamp_from_time_cell(tval, base_dt: datetime) -> str:
    """
    R2（時刻セル）から 'YYYY-MM-DD HH:MM:SS' を作る。無ければ base_dt を使う。
    """
    if tval is None:
        dt = base_dt
    elif isinstance(tval, datetime):
        dt = tval
    elif isinstance(tval, dtime):
        dt = datetime.combine(base_dt.date(), tval)
    else:
        # "HH:MM" or "HH:MM:SS"
        s = str(tval).strip()
        if re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", s):
            if s.count(":") == 1:
                s += ":00"
            hh, mm, ss = map(int, s.split(":"))
            dt = datetime(base_dt.year, base_dt.month, base_dt.day, hh, mm, ss)
        else:
            dt = base_dt
    return dt.strftime("%Y-%m-%d %H:%M:%S")


# ========= スナップショット：セル定義（変えたければここだけ） =========
# 形式: "フィールド名": ("セル番地", 変換関数)
# 変換関数は下の _to_float/_to_int などを利用
SNAPSHOT_MAP: dict[str, tuple[str, Callable[[Any], Any]]] = {
    "last": ("Q2", _to_float),
    "time_cell": ("R2", lambda v: v),
    "diff": ("Q3", _to_float),
    "diff_pct": ("R3", _to_float),
    "high": ("Q4", _to_float),
    "low": ("Q5", _to_float),
    "open": ("Q6", _to_float),
    "prev_close": ("Q7", _to_float),
    "volume": ("U2", _to_int),
    "turnover": ("U3", _to_int),
}


def extract_hex_ticker_from_a1(a1: str) -> str:
    """
    =RssChart(..., "285A.T", ...) から 16進4桁コードを抽出。返り値は大文字。
    """
    m = re.search(r'"([0-9A-Fa-f]{4})\.[Tt]"', a1)
    return m.group(1).upper() if m else ""


def init_db(db_path: Path = DB_PATH) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    for stmt in DDL.strip().split(";"):
        s = stmt.strip()
        if s:
            conn.execute(s)
    conn.commit()
    return conn


# ========= 1分足（既存） =========
def get_last_datetime(conn: sqlite3.Connection, ticker: str) -> str | None:
    cur = conn.execute(
        "SELECT MAX(datetime) FROM minute_data WHERE ticker=?", (ticker,)
    )
    row = cur.fetchone()
    return row[0] if row and row[0] else None


def save_minute_data(
    conn: sqlite3.Connection, ticker: str, sheet_name: str, df: pd.DataFrame
) -> None:
    """DataFrame → minute_data へ差し込み（Pythonの型に正規化）"""
    use = df.copy()
    use["datetime"] = pd.to_datetime(use["datetime"]).dt.strftime("%Y-%m-%d %H:%M:%S")

    rows = []
    for dt, o, h, l, c, v in use[
        ["datetime", "open", "high", "low", "close", "volume"]
    ].itertuples(index=False, name=None):
        rows.append(
            (dt, _to_float(o), _to_float(h), _to_float(l), _to_float(c), _to_int(v))
        )

    if not rows:
        return

    conn.executemany(
        """
        INSERT OR REPLACE INTO minute_data
          (ticker, sheet_name, datetime, open, high, low, close, volume)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [(ticker, sheet_name, *r) for r in rows],
    )
    conn.commit()

    # 直近7営業日だけ残す（ticker単位）
    cur = conn.execute(
        """
        SELECT DISTINCT date(datetime) AS d
        FROM minute_data
        WHERE ticker=?
        ORDER BY d DESC
        LIMIT 7
        """,
        (ticker,),
    )
    keep_days = [r[0] for r in cur.fetchall()]
    if keep_days:
        placeholders = ",".join(["?"] * len(keep_days))
        conn.execute(
            f"DELETE FROM minute_data WHERE ticker=? AND date(datetime) NOT IN ({placeholders})",
            [ticker, *keep_days],
        )
        conn.commit()


def read_excel_fixed(path: Path) -> Dict[Tuple[str, str], pd.DataFrame]:
    """
    固定仕様：
      - 各シート＝1銘柄、シート名＝銘柄名称
      - A1: =RssChart(...,"7453.T",...) から 4桁コードを抽出（ticker）
      - 2行目がヘッダ、3行目以降がデータ
      - 列: A:銘柄名称 B:日付 C:時刻 D:始値 E:高値 F:安値 G:終値 H:出来高
      - B列に '--------' が来たら打ち切り
      - 日付 YYYY/MM/DD、時刻 HH:MM（JST）
      - 無取引行（OHLC全欠損）や volume=0 で OHLC欠損の行は除外
    戻り値: {(ticker, sheet_name): DataFrame(datetime, open, high, low, close, volume)}
    """
    wb = load_workbook(path, data_only=False, read_only=True)
    out: Dict[Tuple[str, str], pd.DataFrame] = {}

    for ws in wb.worksheets:
        sheet_name = ws.title.strip()

        a1 = str(ws["A1"].value or "")
        ticker = extract_hex_ticker_from_a1(a1)
        if not ticker:
            continue

        # 3行目以降（A〜H）
        data_rows = list(
            ws.iter_rows(min_row=3, min_col=1, max_col=8, values_only=True)
        )
        if not data_rows:
            out[(ticker, sheet_name)] = pd.DataFrame(
                columns=["datetime", "open", "high", "low", "close", "volume"]
            )
            continue

        # '--------' で打ち切り（B列=日付）
        cut = None
        for i, r in enumerate(data_rows):
            b = str(r[1]).strip() if r[1] is not None else ""
            if b == "--------":
                cut = i
                break
        if cut is not None:
            data_rows = data_rows[:cut]
            if not data_rows:
                out[(ticker, sheet_name)] = pd.DataFrame(
                    columns=["datetime", "open", "high", "low", "close", "volume"]
                )
                continue

        df = pd.DataFrame(
            data_rows,
            columns=["name", "date", "time", "open", "high", "low", "close", "volume"],
        )
        dt = pd.to_datetime(
            df["date"].astype(str).str.strip()
            + " "
            + df["time"].astype(str).str.strip(),
            format="%Y/%m/%d %H:%M",
            errors="coerce",
        )
        w = pd.DataFrame(
            {
                "datetime": dt,
                "open": pd.to_numeric(df["open"], errors="coerce"),
                "high": pd.to_numeric(df["high"], errors="coerce"),
                "low": pd.to_numeric(df["low"], errors="coerce"),
                "close": pd.to_numeric(df["close"], errors="coerce"),
                "volume": pd.to_numeric(df["volume"], errors="coerce"),
            }
        )

        # 保存時は欠損OHLCや出来高0でも除外せず、ticker/datetimeが正常なら保存
        w = w.dropna(subset=["datetime"])  # datetime不正のみ除外
        w = w.sort_values("datetime").reset_index(drop=True)
        out[(ticker, sheet_name)] = w

    return out


# ========= スナップショット（SNAPSHOT_MAPで抽出） =========
def read_snapshot_with_map(ws: Worksheet) -> dict:
    """SNAPSHOT_MAP に基づきセルを読み、必要な補完を行って dict を返す。"""
    now = datetime.now()
    raw: dict[str, Any] = {}

    for key, (cell, conv) in SNAPSHOT_MAP.items():
        try:
            val = ws[cell].value
            raw[key] = None if val is None else conv(val)
        except Exception:
            raw[key] = None

    updated_at = _timestamp_from_time_cell(raw.get("time_cell"), now)

    snap = {
        "last": raw.get("last"),
        "prev_close": raw.get("prev_close"),
        "open": raw.get("open"),
        "high": raw.get("high"),
        "low": raw.get("low"),
        "volume": raw.get("volume"),
        "turnover": raw.get("turnover"),
        "diff": raw.get("diff"),
        "diff_pct": raw.get("diff_pct"),
        "updated_at": updated_at,
    }

    return snap


def upsert_quote_latest(
    conn: sqlite3.Connection, ticker: str, sheet_name: str, snap: dict
) -> None:
    """quote_latest を UPSERT（ticker ごとに最新1行保持）"""
    conn.execute(
        """
        INSERT INTO quote_latest
          (ticker, sheet_name, last, prev_close, open, high, low, volume, turnover, diff, diff_pct, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(ticker) DO UPDATE SET
          sheet_name=excluded.sheet_name,
          last=excluded.last,
          prev_close=excluded.prev_close,
          open=excluded.open,
          high=excluded.high,
          low=excluded.low,
          volume=excluded.volume,
          turnover=excluded.turnover,
          diff=excluded.diff,
          diff_pct=excluded.diff_pct,
          updated_at=excluded.updated_at
        """,
        (
            ticker,
            sheet_name,
            _to_float(snap.get("last")),
            _to_float(snap.get("prev_close")),
            _to_float(snap.get("open")),
            _to_float(snap.get("high")),
            _to_float(snap.get("low")),
            _to_int(snap.get("volume")),
            _to_int(snap.get("turnover")),
            _to_float(snap.get("diff")),
            _to_float(snap.get("diff_pct")),
            snap.get("updated_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ),
    )
    conn.commit()


def is_marketspeed_running_cmd():
    result = subprocess.run(["tasklist"], capture_output=True, text=True)

    if "marketspeed2.exe" in result.stdout.lower():
        return True
    else:
        return False


# ========= メインループ =========
def main_loop():

    while True:
        loop_start = datetime.now()
        try:
            # 1分足を取得
            data = read_excel_fixed(EXCEL_PATH)

            # スナップショットは data_only=True で1回だけ開く
            wb = load_workbook(EXCEL_PATH, data_only=True, read_only=True)

            for (ticker, sheet_name), df in data.items():
                # --- 1分足：差分投入 ---
                if not df.empty:
                    last_db_dt = get_last_datetime(conn, ticker)
                    if last_db_dt:
                        df["_dt_str"] = pd.to_datetime(df["datetime"]).dt.strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        hit = df.index[df["_dt_str"] == last_db_dt].tolist()
                        start = (hit[0] + 1) if hit else 0
                        df_new = df.iloc[start:].drop(columns=["_dt_str"])
                    else:
                        df_new = df

                    if not df_new.empty:
                        save_minute_data(conn, ticker, sheet_name, df_new)
                        # 最初の行の時刻を取得して表示
                        first_dt = df_new.iloc[0]["datetime"] if "datetime" in df_new.columns else "不明"
                        print(f"[INFO]:{first_dt}  {sheet_name}({ticker}) 追加 {len(df_new)} 行")

                # --- スナップショット：SNAPSHOT_MAP で抽出 & 保存 ---
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    snap = read_snapshot_with_map(ws)
                    upsert_quote_latest(conn, ticker, sheet_name, snap)

            wb.close()

        except Exception as e:
            print(f"[ERROR] {e}")

        # 次の「15秒後」まで待機（処理時間控除）
        next_interval = loop_start + timedelta(seconds=15)
        sleep_time = max(0.0, (next_interval - datetime.now()).total_seconds())
        time.sleep(sleep_time)

if __name__ == "__main__":
    # デイトレ株価データ.xlsmが無ければ即終了
    if not EXCEL_PATH.exists():
        print(f"[INFO] Excelファイルが見つかりません。終了します: {EXCEL_PATH}")
        exit(1)

    # MARKET SPEED2起動確認
    if not is_marketspeed_running_cmd():
        print("⚠️ MARKET SPEED2 が起動していません")
        user_input = input("このまま起動しますか？(Yes/No): ").strip().lower()
        if user_input not in ["yes", "y"]:
            print("プログラムを終了します。")
            exit(2)

    # DB初期化
    conn = init_db(DB_PATH)

    # メインループ開始
    main_loop()
