\"\"\"db_updater_snapshot.py : Auto-generated placeholder

- file: db_updater_snapshot.py
- updated: 2025-09-08

TODO: このモジュールの概要をここに書いてください。
\"\"\"
# db_updater_snapshot.py 〔today_data 監視・置換版 / 15s周期 / 進捗ログ強化〕
from __future__ import annotations
import re, sqlite3, time, subprocess
from datetime import datetime, timedelta, time as dtime
from pathlib import Path
from typing import Any, Callable, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ====== 設定 ======
SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = SCRIPT_DIR / "stock_data.xlsm"      # ← 固定：スナップショットExcel
DB_PATH    = SCRIPT_DIR / "data" / "rss_snapshot.db"
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

# ====== DDL ======
DDL = """
-- 当日（シート内の「最新日付」のみ）を保持する1分足
CREATE TABLE IF NOT EXISTS today_data(
  ticker     TEXT NOT NULL,
  sheet_name TEXT NOT NULL,
  datetime   TEXT NOT NULL,   -- 'YYYY-MM-DD HH:MM:SS' (JST)
  open REAL, high REAL, low REAL, close REAL, volume INTEGER,
  PRIMARY KEY(ticker, datetime)
);
CREATE INDEX IF NOT EXISTS idx_today_t_d ON today_data(ticker, datetime);

-- シート先頭の気配スナップショット（参考）
CREATE TABLE IF NOT EXISTS quote_latest(
  ticker TEXT NOT NULL PRIMARY KEY,
  sheet_name TEXT NOT NULL,
  last REAL, prev_close REAL, open REAL, high REAL, low REAL,
  volume INTEGER, turnover INTEGER, diff REAL, diff_pct REAL,
  updated_at TEXT NOT NULL
);
"""

# ====== ユーティリティ ======
def _to_float(x):
    if x is None: return None
    if isinstance(x, (int, float)): return float(x)
    s = str(x).strip().replace(",", "")
    return float(s) if re.fullmatch(r"-?\d+(\.\d+)?", s) else None

def _to_int(x):
    f = _to_float(x)
    return None if f is None else int(f)

def _timestamp_from_time_cell(tval, base_dt: datetime) -> str:
    if tval is None:
        dt = base_dt
    elif isinstance(tval, datetime):
        dt = tval
    elif isinstance(tval, dtime):
        dt = datetime.combine(base_dt.date(), tval)
    else:
        s = str(tval).strip()
        if re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", s):
            if s.count(":") == 1: s += ":00"
            hh, mm, ss = map(int, s.split(":"))
            dt = datetime(base_dt.year, base_dt.month, base_dt.day, hh, mm, ss)
        else:
            dt = base_dt
    return dt.strftime("%Y-%m-%d %H:%M:%S")

SNAPSHOT_MAP: dict[str, tuple[str, Callable[[Any], Any]]] = {
    "last": ("Q2", _to_float),
    "time_cell": ("R2", lambda v: v),
    "diff": ("Q3", _to_float),
    "diff_pct": ("R3", _to_float),
    "high": ("Q4", _to_float),
    "low": ("Q5", _to_float),
    "open": ("Q6", _to_float),
    "prev_close": ("Q7", _to_float),
    "volume": ("U2", _to_int),
    "turnover": ("U3", _to_int),
}

def extract_ticker_from_a1(a1: str) -> str:
    m = re.search(r'"([0-9A-Fa-f]{4})\.[Tt]"', a1)
    return m.group(1).upper() if m else ""

def init_db() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    for stmt in DDL.strip().split(";"):
        s = stmt.strip()
        if s:
            conn.execute(s)
    conn.commit()
    return conn

def is_marketspeed_running() -> bool:
    r = subprocess.run(["tasklist"], capture_output=True, text=True)
    return "marketspeed2.exe" in r.stdout.lower()

def get_last_datetime(conn: sqlite3.Connection, ticker: str) -> str | None:
    row = conn.execute(
        "SELECT MAX(datetime) FROM today_data WHERE ticker=?", (ticker,)
    ).fetchone()
    return row[0] if row and row[0] else None

def save_today_data(
    conn: sqlite3.Connection, ticker: str, sheet_name: str, df: pd.DataFrame
) -> None:
    use = df.copy()
    use["datetime"] = pd.to_datetime(use["datetime"]).dt.strftime("%Y-%m-%d %H:%M:%S")
    rows = [
        (
            ticker, sheet_name, dt,
            _to_float(o), _to_float(h), _to_float(l), _to_float(c), _to_int(v)
        )
        for dt, o, h, l, c, v in use[
            ["datetime", "open", "high", "low", "close", "volume"]
        ].itertuples(index=False, name=None)
    ]
    if not rows:
        return
    conn.executemany(
        """INSERT OR REPLACE INTO today_data
           (ticker, sheet_name, datetime, open, high, low, close, volume)
           VALUES (?,?,?,?,?,?,?,?)""",
        rows,
    )
    conn.commit()

def read_code_sheet(ws: Worksheet) -> pd.DataFrame:
    # 3行目以降 A〜H： A:銘柄名 B:日付 C:時刻 D:E:F:G:OHLC H:出来高
    rows = list(ws.iter_rows(min_row=3, min_col=1, max_col=8, values_only=True))
    if not rows:
        return pd.DataFrame(columns=["datetime", "open", "high", "low", "close", "volume"])

    # '--------' 手前で打ち切り（B列 = 日付）
    cut = None
    for i, r in enumerate(rows):
        b = str(r[1]).strip() if r[1] is not None else ""
        if b == "--------":
            cut = i
            break
    if cut is not None:
        rows = rows[:cut]
    if not rows:
        return pd.DataFrame(columns=["datetime", "open", "high", "low", "close", "volume"])

    df = pd.DataFrame(rows, columns=["name", "date", "time", "open", "high", "low", "close", "volume"])
    dt = pd.to_datetime(
        df["date"].astype(str).str.strip() + " " + df["time"].astype(str).str.strip(),
        format="%Y/%m/%d %H:%M",
        errors="coerce",
    )
    out = (
        pd.DataFrame({
            "datetime": dt,
            "open":   pd.to_numeric(df["open"],   errors="coerce"),
            "high":   pd.to_numeric(df["high"],   errors="coerce"),
            "low":    pd.to_numeric(df["low"],    errors="coerce"),
            "close":  pd.to_numeric(df["close"],  errors="coerce"),
            "volume": pd.to_numeric(df["volume"], errors="coerce"),
        })
        .dropna(subset=["datetime"])           # 時刻不明は除外
        .sort_values("datetime")
        .reset_index(drop=True)
    )
    return out

def latest_day_only(df: pd.DataFrame) -> pd.DataFrame:
    """シート内に混在しても「最新日付のみ」を返す"""
    if df.empty:
        return df
    dts = pd.to_datetime(df["datetime"])
    last_date = dts.dt.date.max()
    return df[dts.dt.date == last_date].copy()

def read_snapshot(ws: Worksheet) -> dict:
    now = datetime.now()
    raw = {}
    for k, (cell, conv) in SNAPSHOT_MAP.items():
        try:
            v = ws[cell].value
            raw[k] = None if v is None else conv(v)
        except Exception:
            raw[k] = None
    return {
        "last":       raw.get("last"),
        "prev_close": raw.get("prev_close"),
        "open":       raw.get("open"),
        "high":       raw.get("high"),
        "low":        raw.get("low"),
        "volume":     raw.get("volume"),
        "turnover":   raw.get("turnover"),
        "diff":       raw.get("diff"),
        "diff_pct":   raw.get("diff_pct"),
        "updated_at": _timestamp_from_time_cell(raw.get("time_cell"), now),
    }

def upsert_quote_latest(conn: sqlite3.Connection, ticker: str, sheet_name: str, snap: dict) -> None:
    conn.execute(
        """
        INSERT INTO quote_latest
          (ticker, sheet_name, last, prev_close, open, high, low, volume, turnover, diff, diff_pct, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(ticker) DO UPDATE SET
          sheet_name=excluded.sheet_name,
          last=excluded.last,
          prev_close=excluded.prev_close,
          open=excluded.open,
          high=excluded.high,
          low=excluded.low,
          volume=excluded.volume,
          turnover=excluded.turnover,
          diff=excluded.diff,
          diff_pct=excluded.diff_pct,
          updated_at=excluded.updated_at
        """,
        (
            ticker, sheet_name,
            _to_float(snap.get("last")), _to_float(snap.get("prev_close")),
            _to_float(snap.get("open")), _to_float(snap.get("high")), _to_float(snap.get("low")),
            _to_int(snap.get("volume")), _to_int(snap.get("turnover")),
            _to_float(snap.get("diff")), _to_float(snap.get("diff_pct")),
            snap.get("updated_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ),
    )
    conn.commit()

def get_targets(wb) -> List[Tuple[str, str]]:
    """indexシート A1〜A30 に 4桁コード、対応する code1..code30 シートを対象とする"""
    if "index" not in wb.sheetnames:
        return []
    ws = wb["index"]
    targets: List[Tuple[str, str]] = []
    for i in range(1, 31):
        val = ws[f"A{i}"].value
        code = str(val).strip() if val is not None else ""
        if re.fullmatch(r"[0-9A-Fa-f]{4}", code or ""):
            sname = f"code{i}"
            if sname in wb.sheetnames:
                targets.append((code.upper(), sname))
    return targets

def cleanup_to_latest_date(conn: sqlite3.Connection, ticker: str) -> None:
    """保険：ticker単位で『最新日付のみ』をDBに残す（混入防止）"""
    row = conn.execute(
        "SELECT MAX(date(datetime)) FROM today_data WHERE ticker=?", (ticker,)
    ).fetchone()
    if not row or not row[0]:
        return
    latest_date = row[0]  # 'YYYY-MM-DD'
    conn.execute(
        "DELETE FROM today_data WHERE ticker=? AND date(datetime)<>?",
        (ticker, latest_date),
    )
    conn.commit()

def clear_today_data(conn: sqlite3.Connection):
    """起動時に today_data をクリア"""
    conn.execute("DELETE FROM today_data")
    conn.commit()
    print("[INIT] today_data をクリアしました")

def main_loop(conn: sqlite3.Connection):
    print(f"[START] db_updater_snapshot.py")
    print(f"[CONFIG] EXCEL={EXCEL_PATH.name}  DB={DB_PATH}")

    while True:
        t0 = datetime.now()
        try:
            # 値セル版（データ混在検知に使う）＋ data_only 版（スナップショット用）
            wb_values = load_workbook(EXCEL_PATH, data_only=False, read_only=True)
            wb_snap   = load_workbook(EXCEL_PATH, data_only=True,  read_only=True)
            targets = get_targets(wb_values)
            if not targets:
                print("[INFO] indexシートに対象がありません")
            for ticker, sheet_name in targets:
                ws = wb_values[sheet_name]

                # A1 の "XXXX.T" があれば上書き（保険）
                a1 = str(ws["A1"].value or "")
                tk_from_a1 = extract_ticker_from_a1(a1)
                if re.fullmatch(r"[0-9A-Fa-f]{4}", tk_from_a1 or ""):
                    ticker = tk_from_a1.upper()

                # シート → DataFrame（'--------'手前まで）→ 最新日付だけ
                df = read_code_sheet(ws)
                df = latest_day_only(df)

                if not df.empty:
                    # 既存の直近日時より新しい分だけ追加
                    last_dt = get_last_datetime(conn, ticker)
                    if last_dt:
                        s = pd.to_datetime(df["datetime"]).dt.strftime("%Y-%m-%d %H:%M:%S")
                        newer_mask = s > last_dt
                        df_new = df.loc[newer_mask].copy()
                    else:
                        df_new = df

                    if not df_new.empty:
                        save_today_data(conn, ticker, sheet_name, df_new)
                        first_dt = df_new.iloc[0]["datetime"]
                        print(f"[INFO]:{first_dt}  {sheet_name}({ticker}) 追加 {len(df_new)} 行")

                    # 念のため『最新日付のみ残す』をDB側でも実施
                    cleanup_to_latest_date(conn, ticker)

                # スナップショット更新（data_only=True）
                if sheet_name in wb_snap.sheetnames:
                    upsert_quote_latest(conn, ticker, sheet_name, read_snapshot(wb_snap[sheet_name]))

            wb_values.close()
            wb_snap.close()
        except Exception as e:
            print(f"[ERROR] {e}")

        # 次の「15秒後」まで待機（処理時間控除）
        next_t = t0 + timedelta(seconds=15)
        sleep_time = max(0.0, (next_t - datetime.now()).total_seconds())
        print(f"[LOOP] {datetime.now().strftime('%H:%M:%S')} 完了（sleep {sleep_time:.1f}s）")
        time.sleep(sleep_time)

if __name__ == "__main__":
    if not EXCEL_PATH.exists():
        print(f"[INFO] Excelが見つかりません: {EXCEL_PATH}")
        raise SystemExit(1)

    # MarketSpeed2 起動確認（必要なければコメントアウト可）
    if not is_marketspeed_running():
        print("⚠️ MARKET SPEED2 が起動していません")
        ans = input("続行しますか？(Yes/No): ").strip().lower()
        if ans not in ("y", "yes"):
            raise SystemExit(2)

    conn = init_db()
    clear_today_data(conn)
    main_loop(conn)
