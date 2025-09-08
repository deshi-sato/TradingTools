\"\"\"excel_loader.py : Auto-generated placeholder

- file: excel_loader.py
- updated: 2025-09-08

TODO: このモジュールの概要をここに書いてください。
\"\"\"
import openpyxl
from score_table import create_score_table_long, create_score_table_short
from tqdm import tqdm
import pandas as pd
import shutil
import xlwings as xw
from utility import parse_date_time

EXCEL_PATH_L = "C:/Users/Owner/Documents/desshi_signal_viewer/買い銘柄寄り後情報.xlsm"
EXCEL_PATH_S = "C:/Users/Owner/Documents/desshi_signal_viewer/売り銘柄寄り後情報.xlsm"

RSS_PARAM_TO_REPLACE = "1660"
RSS_PARAM_NEW = "332"
SCORE_THRESHOLD_L = 7
SCORE_THRESHOLD_S = 4


def get_latest_date_from_data(file_path):
    """Excelファイルから最新の日付を取得する"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheetnames = wb.sheetnames
    latest_date = None

    for sheet_name in sheetnames[:5]:  # 最初の5シートで確認
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=3, values_only=True):
            # データ終端のチェック
            if isinstance(row[1], str) and "----" in str(row[1]):
                break

            if (
                row[1] is None
                or row[2] is None
                or row[3] is None
                or row[4] is None
                or row[5] is None
                or row[7] == 0
            ):
                continue

            try:
                dt = parse_date_time(row[1], row[2])
                date_key = dt.strftime("%Y-%m-%d")

                if latest_date is None or date_key > latest_date:
                    latest_date = date_key

            except Exception as e:
                continue

    wb.close()
    return latest_date


def load_summary_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheetnames = wb.sheetnames
    data_dict = {}
    code_to_name = {}  # ← 銘柄コード→名称の対応辞書を追加

    # ✅ A1の値を確認し、RSS通信が未確立なら中断
    first_sheet = wb[sheetnames[0]]
    a1_value = str(first_sheet["A1"].value)
    if "#NAME?" in a1_value or a1_value.strip() == "":
        print(f"⚠️ 通信未確立（A1セル = {a1_value}）のため読み込み中断: {file_path}")
        return {}, {}

    # 最新の日付を取得
    latest_date = get_latest_date_from_data(file_path)
    if latest_date is None:
        print(f"⚠️ データから最新日付を取得できませんでした: {file_path}")
        return {}, {}

    print(f"📅 最新日付: {latest_date}")

    for sheet_name in tqdm(sheetnames, desc="Excel読み込み中"):
        ws = wb[sheet_name]

        # ✅ A1から銘柄コード抽出（例: "5803.T" → "5803"）
        try:
            formula = str(ws["A1"].value)
            code = formula.split(",")[1].strip().strip('"').split(".")[0]
        except Exception as e:
            print(
                f"❌ シート「{sheet_name}」のA1({formula})から銘柄コード抽出失敗: {e}"
            )
            continue

        records = []

        for row in ws.iter_rows(min_row=3, values_only=True):
            # データ終端のチェック
            if isinstance(row[1], str) and "----" in str(row[1]):
                break

            if (
                row[1] is None
                or row[2] is None
                or row[3] is None
                or row[4] is None
                or row[5] is None
                or row[7] == 0
            ):
                continue

            try:
                dt = parse_date_time(row[1], row[2])
                date_key = dt.strftime("%Y-%m-%d")
                if date_key != latest_date:
                    continue  # 最新日付以外は除外

                record = {
                    "time": dt,
                    "open": row[3],
                    "high": row[4],
                    "low": row[5],
                    "close": row[6],
                    "volume": row[7],
                }
                records.append(record)
            except Exception as e:
                print(f"{sheet_name} の行でエラー: {e}")
                continue

        if records:
            df = pd.DataFrame(records)
            data_dict[code] = {latest_date: df}
            code_to_name[code] = sheet_name
            print(
                f"✅ コード {code} ← シート「{sheet_name}」最新日 {latest_date} {len(records)}本"
            )

    return data_dict, code_to_name  # ← 2つ返す


def load_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheetnames = wb.sheetnames
    data_dict = {}
    code_to_name = {}  # ← 銘柄コード→名称の対応辞書を追加

    for sheet_name in tqdm(sheetnames, desc="Excel読み込み中"):
        ws = wb[sheet_name]

        # ✅ A1のRSS関数から銘柄コードを抽出
        try:
            formula = str(ws["A1"].value)
            code = (
                formula.split(",")[1].strip().strip('"').split(".")[0]
            )  # "5803.T" → "5803"
        except Exception as e:
            print(f"❌ シート「{sheet_name}」のA1から銘柄コード抽出失敗: {e}")
            continue

        daily_rows = {}

        for row in ws.iter_rows(min_row=3, values_only=True):
            # データ終端のチェック
            if isinstance(row[1], str) and "----" in str(row[1]):
                break

            if (
                row[1] is None
                or row[2] is None
                or row[3] is None
                or row[4] is None
                or row[5] is None
                or row[7] == 0
            ):
                continue

            try:
                dt = parse_date_time(row[1], row[2])
                record = {
                    "time": dt,
                    "open": row[3],
                    "high": row[4],
                    "low": row[5],
                    "close": row[6],
                    "volume": row[7],
                }
                date_key = dt.strftime("%Y-%m-%d")
                if date_key not in daily_rows:
                    daily_rows[date_key] = []
                daily_rows[date_key].append(record)
            except Exception as e:
                print(f"{sheet_name} の行でエラー: {e}")
                continue

        daily_frames = {
            day: pd.DataFrame(records)
            for day, records in daily_rows.items()
            if len(records) >= 300
        }

        if len(daily_frames) >= 3:
            data_dict[code] = daily_frames
            code_to_name[code] = sheet_name  # ← 対応を登録
    #            print(
    #                f"✅ コード {code} ← シート「{sheet_name}」として登録（{len(daily_frames)}日分）"
    #            )

    return data_dict, code_to_name  # ← 2つ返す


def export_sheets(src_path, top_long, top_short, code_to_name):
    global EXCEL_PATH_L, EXCEL_PATH_S

    def process_copy(dst_path, code_list):

        shutil.copy(src_path, dst_path)
        app = xw.App(visible=False)
        app.display_alerts = False
        wb = app.books.open(dst_path)

        # コードからシート名（銘柄名）に変換
        sheet_name_list = [code_to_name.get(code, "") for code in code_list]

        for sheet in tqdm(wb.sheets, desc="シート削除中"):
            if sheet.name not in sheet_name_list:
                try:
                    sheet.delete()
                except Exception as e:
                    print(f"⚠️ シート {sheet.name} の削除に失敗: {e}")
            else:
                formula = sheet.range("A1").formula
                if (
                    isinstance(formula, str)
                    and formula.startswith("=RssChart")
                    and f", {RSS_PARAM_TO_REPLACE}" in formula
                ):
                    sheet.range("A1").formula = formula.replace(
                        f", {RSS_PARAM_TO_REPLACE}", f", {RSS_PARAM_NEW}"
                    )

        wb.save()
        wb.close()
        app.quit()

    if top_long is not None and not top_long.empty:
        print(f"📊 買いスコア上位:\n{top_long}")
        process_copy(EXCEL_PATH_L, top_long["ticker"].tolist())

    if top_short is not None and not top_short.empty:
        print(f"📉 売りスコア上位:\n{top_short}")
        process_copy(EXCEL_PATH_S, top_short["ticker"].tolist())


def export_top_sheets():
    src_path = "デイトレ株価データ.xlsx"

    # スコア表読み込み
    long_df = pd.read_csv("score_table_long.csv", encoding="shift_jis")
    short_df = pd.read_csv("score_table_short.csv", encoding="shift_jis")

    # スコア7点以上のみ抽出
    top_long = long_df[long_df["合計スコア"] >= SCORE_THRESHOLD_L]["ticker"].tolist()
    top_short = short_df[short_df["合計スコア"] >= SCORE_THRESHOLD_S]["ticker"].tolist()

    # Excel起動
    app = xw.App(visible=False)
    wb_src = app.books.open(src_path)

    # ✅ 買い銘柄ファイルの作成
    if top_long:
        wb_long = app.books.add()
        for sheet_name in top_long:
            if sheet_name in [s.name for s in wb_src.sheets]:
                wb_src.sheets[sheet_name].copy(after=wb_long.sheets[-1])
            else:
                print(f"⚠️ 買いシート {sheet_name} が見つかりません")
        if len(wb_long.sheets) > 1 and wb_long.sheets[0].name == "Sheet1":
            wb_long.sheets[0].delete()
        for sheet in wb_long.sheets:
            formula = sheet.range("A1").formula
            if (
                isinstance(formula, str)
                and formula.startswith("=RssChart")
                and f", {RSS_PARAM_TO_REPLACE}" in formula
            ):
                sheet.range("A1").formula = formula.replace(
                    f", {RSS_PARAM_TO_REPLACE}", f", {RSS_PARAM_NEW}"
                )
        wb_long.save("買い銘柄寄り後情報.xlsx")
        wb_long.close()

    # ✅ 売り銘柄ファイルの作成
    if top_short:
        wb_short = app.books.add()
        for sheet_name in top_short:
            if sheet_name in [s.name for s in wb_src.sheets]:
                wb_src.sheets[sheet_name].copy(after=wb_short.sheets[-1])
            else:
                print(f"⚠️ 売りシート {sheet_name} が見つかりません")
        if len(wb_short.sheets) > 1 and wb_short.sheets[0].name == "Sheet1":
            wb_short.sheets[0].delete()
        for sheet in wb_short.sheets:
            formula = sheet.range("A1").formula
            if (
                isinstance(formula, str)
                and formula.startswith("=RssChart")
                and f", {RSS_PARAM_TO_REPLACE}" in formula
            ):
                sheet.range("A1").formula = formula.replace(
                    f", {RSS_PARAM_TO_REPLACE}", f", {RSS_PARAM_NEW}"
                )
        wb_short.save("売り銘柄寄り後情報.xlsx")
        wb_short.close()

    wb_src.close()
    app.quit()


# 実行
if __name__ == "__main__":
    print("デイトレ株価データからスコア表を作成します")
    excel_path = "デイトレ株価データ.xlsx"
    data_dict = load_data(excel_path)

    # 買い候補（ロング）
    result_long = create_score_table_long(data_dict)
    result_long.to_csv("score_table_long.csv", index=False, encoding="shift_jis")

    # 売り候補（ショート）
    result_short = create_score_table_short(data_dict)
    result_short.to_csv("score_table_short.csv", index=False, encoding="shift_jis")

    export_top_sheets()
    print("スコア表作成完了")
