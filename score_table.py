#
# 　デイトレ用　推奨銘柄スコア表作成
# 　Ver 1.25.7.25
#
# 　入力：同一フォルダにあるデイトレ株価データ.xlsm
# 　出力：score_table_long.csv（買い目線スコア表）
# 　　　：score_table_short.csv（売り目線スコア表）
# 　　　：推薦銘柄寄り後情報.xlsx（買い・売り最終候補を寄り15分後のチェック用）
#
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from tqdm import tqdm
import xlwings as xw
import os
import configparser
import shutil
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from flask import Flask, render_template
import mplfinance as mpf

app = Flask(__name__)

# フォント設定（Noto Sans CJK JP を使用）
plt.rcParams["font.family"] = "Yu Gothic"

EXCEL_PATH = "C:/Users/Owner/Documents/desshi_signal_viewer/デイトレ株価データ.xlsm"
EXCEL_PATH_L = "C:/Users/Owner/Documents/desshi_signal_viewer/買い銘柄寄り後情報.xlsm"
EXCEL_PATH_S = "C:/Users/Owner/Documents/desshi_signal_viewer/売り銘柄寄り後情報.xlsm"
TEMP_PATH = "C:/Users/Owner/Documents/desshi_signal_viewer/temp_デイトレ株価データ.xlsm"

SCORE_THRESHOLD_L = 7
SCORE_THRESHOLD_S = 4
RSS_PARAM_TO_REPLACE = "1660"
RSS_PARAM_NEW = "332"

# === .ini 管理設定 ===
INI_PATH = "desshi_signal_viewer.ini"


def get_latest_row_index():
    config = configparser.ConfigParser()
    config.read(INI_PATH)
    try:
        return int(config["読み込み状態"]["latest_row_index"])
    except:
        return 0


def save_latest_row_index(index):
    config = configparser.ConfigParser()
    if not os.path.exists(INI_PATH):
        config["読み込み状態"] = {"latest_row_index": str(index)}
    else:
        config.read(INI_PATH)
        config["読み込み状態"]["latest_row_index"] = str(index)
    with open(INI_PATH, "w") as f:
        config.write(f)


def parse_date_time(row_date, row_time):
    if isinstance(row_date, str):
        row_date = pd.to_datetime(row_date).date()
    elif isinstance(row_date, datetime):
        row_date = row_date.date()
    if isinstance(row_time, str):
        row_time = pd.to_datetime(row_time).time()
    elif isinstance(row_time, datetime):
        row_time = row_time.time()
    return datetime.combine(row_date, row_time)


def get_japan_market_today():
    now = datetime.now()
    market_start = now.replace(hour=9, minute=0, second=0, microsecond=0)
    if now < market_start:
        # 9:00より前 → 前日を「今日」とする
        return (now - timedelta(days=1)).strftime("%Y-%m-%d")
    else:
        # 9:00以降 → 通常の今日
        return now.strftime("%Y-%m-%d")


def get_latest_date_from_data(file_path):
    """Excelファイルから最新の日付を取得する"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheetnames = wb.sheetnames
    latest_date = None

    for sheet_name in sheetnames[:5]:  # 最初の5シートで確認
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=3, values_only=True):
            # データ終端のチェック
            if isinstance(row[1], str) and "----" in str(row[1]):
                break

            if (
                row[1] is None
                or row[2] is None
                or row[3] is None
                or row[4] is None
                or row[5] is None
                or row[7] == 0
            ):
                continue

            try:
                dt = parse_date_time(row[1], row[2])
                date_key = dt.strftime("%Y-%m-%d")

                if latest_date is None or date_key > latest_date:
                    latest_date = date_key

            except Exception as e:
                continue

    wb.close()
    return latest_date


def load_summary_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheetnames = wb.sheetnames
    data_dict = {}
    code_to_name = {}  # ← 銘柄コード→名称の対応辞書を追加

    # ✅ A1の値を確認し、RSS通信が未確立なら中断
    first_sheet = wb[sheetnames[0]]
    a1_value = str(first_sheet["A1"].value)
    if "#NAME?" in a1_value or a1_value.strip() == "":
        print(f"⚠️ 通信未確立（A1セル = {a1_value}）のため読み込み中断: {file_path}")
        return {}, {}

    # 最新の日付を取得
    latest_date = get_latest_date_from_data(file_path)
    if latest_date is None:
        print(f"⚠️ データから最新日付を取得できませんでした: {file_path}")
        return {}, {}

    print(f"📅 最新日付: {latest_date}")

    for sheet_name in tqdm(sheetnames, desc="Excel読み込み中"):
        ws = wb[sheet_name]

        # ✅ A1から銘柄コード抽出（例: "5803.T" → "5803"）
        try:
            formula = str(ws["A1"].value)
            code = formula.split(",")[1].strip().strip('"').split(".")[0]
        except Exception as e:
            print(
                f"❌ シート「{sheet_name}」のA1({formula})から銘柄コード抽出失敗: {e}"
            )
            continue

        records = []

        for row in ws.iter_rows(min_row=3, values_only=True):
            # データ終端のチェック
            if isinstance(row[1], str) and "----" in str(row[1]):
                break

            if (
                row[1] is None
                or row[2] is None
                or row[3] is None
                or row[4] is None
                or row[5] is None
                or row[7] == 0
            ):
                continue

            try:
                dt = parse_date_time(row[1], row[2])
                date_key = dt.strftime("%Y-%m-%d")
                if date_key != latest_date:
                    continue  # 最新日付以外は除外

                record = {
                    "time": dt,
                    "open": row[3],
                    "high": row[4],
                    "low": row[5],
                    "close": row[6],
                    "volume": row[7],
                }
                records.append(record)
            except Exception as e:
                print(f"{sheet_name} の行でエラー: {e}")
                continue

        if records:
            df = pd.DataFrame(records)
            data_dict[code] = {latest_date: df}
            code_to_name[code] = sheet_name
            print(
                f"✅ コード {code} ← シート「{sheet_name}」最新日 {latest_date} {len(records)}本"
            )

    return data_dict, code_to_name  # ← 2つ返す


def save_chart_5min(ticker, df, global_data_dict):
    import matplotlib.pyplot as plt
    import mplfinance as mpf
    from datetime import timedelta

    prev_open = prev_high = prev_low = prev_close = None

    if ticker not in global_data_dict:
        print("global_data_dictが不正または空です。チャート作成スキップ。")
        return None

    # 5分足に変換
    df_resampled = (
        df.resample("5min", on="time")
        .agg(
            {
                "open": "first",
                "high": "max",
                "low": "min",
                "close": "last",
                "volume": "sum",
            }
        )
        .dropna()
    )

    # ✅ resample後チェック
    if (
        df_resampled.empty
        or df_resampled[["open", "high", "low", "close"]].dropna().empty
    ):
        print(f"⚠️ {ticker} の5分足データが不正または空です。チャート作成スキップ。")
        print(f"   df_resampled.shape: {df_resampled.shape}")
        print(f"   df_resampled.columns: {df_resampled.columns.tolist()}")
        return None

    df_resampled.index.name = "Date"
    df_resampled.reset_index(inplace=True)
    df_resampled.set_index("Date", inplace=True)

    # テクニカル指標
    df_resampled["5MA"] = df_resampled["close"].rolling(window=5).mean()
    df_resampled["25MA"] = df_resampled["close"].rolling(window=25).mean()
    df_resampled["VWAP"] = (
        df_resampled["close"] * df_resampled["volume"]
    ).cumsum() / df_resampled["volume"].cumsum()

    # カラム名変換
    df_plot = df_resampled.rename(
        columns={
            "open": "Open",
            "high": "High",
            "low": "Low",
            "close": "Close",
            "volume": "Volume",
        }
    )
    line_len = len(df_plot)

    # 最終チェック
    if (
        df_plot.empty
        or df_plot[["Open", "High", "Low", "Close"]].isnull().all().any()
        or df_plot[["Open", "High", "Low", "Close"]].dropna().shape[0] < 3
    ):
        print(f"⚠️ {ticker} の描画データが不正（空 or NaN or 3本未満）。スキップ。")
        print(f"   df_plot.shape: {df_plot.shape}")
        print(f"   df_plot.columns: {df_plot.columns.tolist()}")
        print(
            f"   df_plot[['Open', 'High', 'Low', 'Close']].dropna().shape: {df_plot[['Open', 'High', 'Low', 'Close']].dropna().shape}"
        )
        return None

    # インジケーターがNaNだけでないかチェック関数
    def is_valid_series(s, min_count=3):
        return s.dropna().shape[0] >= min_count

    # 安全なインジケーターだけを描画に追加
    add_plots = []

    # 当日の日付（df は当日分だけ）
    today = df["time"].dt.date.iloc[0]
    yesterday_str = str(today - timedelta(days=1))

    # global_data_dictから最新日付の前日を取得
    daily_dict = global_data_dict[ticker]
    if isinstance(daily_dict, dict) and daily_dict:
        # 日付を降順でソートして最新日付を取得
        sorted_dates = sorted(daily_dict.keys(), reverse=True)
        if sorted_dates:
            yesterday_str = sorted_dates[1]

    # グローバルから該当データ取得
    prev_df = global_data_dict.get(ticker, {}).get(yesterday_str)

    # 🔽 当日（df_plot）の範囲を取得
    today_high = df_plot["High"].max()
    today_low = df_plot["Low"].min()

    # 🔽 前日データを取得
    daily_dict = global_data_dict[ticker]
    if not isinstance(daily_dict, dict):
        print(f"⚠️ {ticker} に対応する値が dict ではありません: {type(daily_dict)}")
        return None
    elif yesterday_str not in daily_dict:
        print(f"⚠️ {ticker} は存在するが {yesterday_str} のデータがありません")
        return None
    else:
        prev_df = daily_dict[yesterday_str]

        # 前日データが存在する場合のみ前日四本値を取得
        if not prev_df.empty:
            try:
                prev_open = prev_df["open"].iloc[0]
                prev_high = prev_df["high"].max()
                prev_low = prev_df["low"].min()
                prev_close = prev_df["close"].iloc[-1]

                # 値がNoneまたはNaNでないことを確認
                if (
                    pd.isna(prev_open)
                    or pd.isna(prev_high)
                    or pd.isna(prev_low)
                    or pd.isna(prev_close)
                ):
                    prev_open = prev_high = prev_low = prev_close = None
            except Exception as e:
                prev_open = prev_high = prev_low = prev_close = None

            # 🔽 チャート範囲に含まれるOHLCのみライン追加
            if (
                prev_open is not None
                and isinstance(prev_open, (int, float))
                and not pd.isna(prev_open)
                and today_low <= prev_open <= today_high
            ):
                try:
                    add_plots.append(
                        mpf.make_addplot(
                            [float(prev_open)] * line_len,
                            panel=0,
                            color="gray",
                            linestyle="--",
                            width=0.8,
                        )
                    )
                except (ValueError, TypeError) as e:
                    pass

            if (
                prev_close is not None
                and isinstance(prev_close, (int, float))
                and not pd.isna(prev_close)
                and today_low <= prev_close <= today_high
            ):
                try:
                    add_plots.append(
                        mpf.make_addplot(
                            [float(prev_close)] * line_len,
                            panel=0,
                            color="black",
                            linestyle="--",
                            width=0.8,
                        )
                    )
                except (ValueError, TypeError) as e:
                    pass

            if (
                prev_high is not None
                and isinstance(prev_high, (int, float))
                and not pd.isna(prev_high)
                and today_low <= prev_high <= today_high
            ):
                try:
                    add_plots.append(
                        mpf.make_addplot(
                            [float(prev_high)] * line_len,
                            panel=0,
                            color="red",
                            linestyle=":",
                            width=0.8,
                        )
                    )
                except (ValueError, TypeError) as e:
                    pass

            if (
                prev_low is not None
                and isinstance(prev_low, (int, float))
                and not pd.isna(prev_low)
                and today_low <= prev_low <= today_high
            ):
                try:
                    add_plots.append(
                        mpf.make_addplot(
                            [float(prev_low)] * line_len,
                            panel=0,
                            color="blue",
                            linestyle=":",
                            width=0.8,
                        )
                    )
                except (ValueError, TypeError) as e:
                    pass

    if is_valid_series(df_plot["VWAP"]):
        try:
            # VWAPデータが数値型であることを確認
            vwap_data = df_plot["VWAP"].dropna()
            if not vwap_data.empty and vwap_data.dtype in ["float64", "int64"]:
                add_plots.append(
                    mpf.make_addplot(df_plot["VWAP"], color="orange", linestyle="-.")
                )
        except Exception as e:
            pass

    path = f"static/chart_{ticker}_5min.png"

    #    print(f"✅ チャート描画直前: {ticker}")
    #    print(df_plot.tail())
    #    print(df_plot[["Open", "High", "Low", "Close"]].info())

    try:
        s = mpf.make_mpf_style(
            # 基本はdefaultの設定値を使う。
            base_mpf_style="default",
            # font.family を matplotlibに設定されている値にする。
            rc={"font.family": plt.rcParams["font.family"][0]},
        )
        fig, axes = mpf.plot(
            df_plot,
            type="candle",
            mav=(5, 25),
            style=s,
            addplot=add_plots,
            ylabel="株価",
            ylabel_lower="出来高",
            volume=True,
            figsize=(20, 6),
            returnfig=True,  # ← fig, axes を取得する
        )

        # 🔽 前日四本値を注釈として下に表示
        if (
            prev_df is not None
            and not prev_df.empty
            and "prev_open" in locals()
            and prev_open is not None
            and prev_high is not None
            and prev_low is not None
            and prev_close is not None
            and isinstance(prev_open, (int, float))
            and isinstance(prev_high, (int, float))
            and isinstance(prev_low, (int, float))
            and isinstance(prev_close, (int, float))
            and not pd.isna(prev_open)
            and not pd.isna(prev_high)
            and not pd.isna(prev_low)
            and not pd.isna(prev_close)
        ):
            try:
                text_str = (
                    f"前日 始: {float(prev_open):.2f}  高: {float(prev_high):.2f}  "
                    f"安: {float(prev_low):.2f}  終: {float(prev_close):.2f}"
                )
                axes[0].text(
                    0.01,
                    -0.18,  # 左下の少し下
                    text_str,
                    transform=axes[0].transAxes,
                    fontsize=10,
                    verticalalignment="top",
                )
            except (ValueError, TypeError) as e:
                pass

        # staticディレクトリが存在しない場合は作成
        import os

        os.makedirs("static", exist_ok=True)

        fig.savefig(path, dpi=200, bbox_inches="tight")
        plt.close(fig)

        # ファイルが実際に保存されたかチェック
        if not os.path.exists(path):
            print(f"❌ {ticker} チャートファイルが保存されませんでした: {path}")
            return None

    except Exception as e:
        print(f"❌ {ticker} チャート描画失敗: {e}")
        return None

    return path


def load_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheetnames = wb.sheetnames
    data_dict = {}
    code_to_name = {}  # ← 銘柄コード→名称の対応辞書を追加

    for sheet_name in tqdm(sheetnames, desc="Excel読み込み中"):
        ws = wb[sheet_name]

        # ✅ A1のRSS関数から銘柄コードを抽出
        try:
            formula = str(ws["A1"].value)
            code = (
                formula.split(",")[1].strip().strip('"').split(".")[0]
            )  # "5803.T" → "5803"
        except Exception as e:
            print(f"❌ シート「{sheet_name}」のA1から銘柄コード抽出失敗: {e}")
            continue

        daily_rows = {}

        for row in ws.iter_rows(min_row=3, values_only=True):
            # データ終端のチェック
            if isinstance(row[1], str) and "----" in str(row[1]):
                break

            if (
                row[1] is None
                or row[2] is None
                or row[3] is None
                or row[4] is None
                or row[5] is None
                or row[7] == 0
            ):
                continue

            try:
                dt = parse_date_time(row[1], row[2])
                record = {
                    "time": dt,
                    "open": row[3],
                    "high": row[4],
                    "low": row[5],
                    "close": row[6],
                    "volume": row[7],
                }
                date_key = dt.strftime("%Y-%m-%d")
                if date_key not in daily_rows:
                    daily_rows[date_key] = []
                daily_rows[date_key].append(record)
            except Exception as e:
                print(f"{sheet_name} の行でエラー: {e}")
                continue

        daily_frames = {
            day: pd.DataFrame(records)
            for day, records in daily_rows.items()
            if len(records) >= 300
        }

        if len(daily_frames) >= 3:
            data_dict[code] = daily_frames
            code_to_name[code] = sheet_name  # ← 対応を登録
    #            print(
    #                f"✅ コード {code} ← シート「{sheet_name}」として登録（{len(daily_frames)}日分）"
    #            )

    return data_dict, code_to_name  # ← 2つ返す


def evaluate_stock_long(day_frames):
    score = {
        "trend": 0,
        "volume": 0,
        "break": 0,
        "close_pos": 0,
        "volatility": 0,
        "vol_level": 0,
    }

    if len(day_frames) < 3:
        return score  # 評価に必要な最低日数に満たない

    # 日付の降順で最新5日を抽出
    sorted_days = sorted(day_frames.keys(), reverse=True)
    recent_days = sorted_days[:5]
    frames = [day_frames[day] for day in recent_days if day in day_frames]

    # 日別の終値・高値・安値・出来高を取得
    closes = [df["close"].iloc[-1] for df in frames if not df.empty]
    highs = [df["high"].max() for df in frames if not df.empty]
    lows = [df["low"].min() for df in frames if not df.empty]
    volumes = [df["volume"].sum() for df in frames if not df.empty]

    # ①トレンド
    if len(highs) >= 3 and len(lows) >= 3:
        if highs[2] < highs[1] < highs[0] and lows[2] < lows[1] < lows[0]:
            score["trend"] = 2
        elif highs[1] < highs[0] or lows[1] < lows[0]:
            score["trend"] = 1

    # ②出来高変化（直近 vs 前日）
    if len(volumes) >= 2 and volumes[1] > 0:
        ratio = (volumes[0] - volumes[1]) / volumes[1]
        if ratio >= 0.2:
            score["volume"] = 2
        elif ratio >= 0.05:
            score["volume"] = 1

    # ③ブレイク位置（終値 vs 前日高値）
    if len(highs) >= 2 and len(closes) >= 1:
        diff = (closes[0] - highs[1]) / highs[1]
        if diff >= 0.005:
            score["break"] = 2
        elif abs(diff) < 0.005:
            score["break"] = 1

    # ④引け位置（終値が当日高値に近い）
    today_high = highs[0]
    today_close = closes[0]
    if today_high:
        diff = (today_high - today_close) / today_high
        if diff <= 0.005:
            score["close_pos"] = 2
        elif diff <= 0.01:
            score["close_pos"] = 1

    # ⑤ボラティリティ（当日）
    today_low = lows[0]
    if today_close > 0 and (today_high - today_low) / today_close >= 0.03:
        score["volatility"] = 1

    # ⑥出来高水準（過去4日平均と比べて1.5倍以上）
    if len(volumes) >= 5:
        avg_volume = sum(volumes[1:5]) / 4
        if volumes[0] >= avg_volume * 1.5:
            score["vol_level"] = 1

    return score


def evaluate_stock_short(day_frames):
    score = {
        "trend": 0,
        "volume": 0,
        "break": 0,
        "close_pos": 0,
        "volatility": 0,
        "vol_level": 0,
    }

    if len(day_frames) < 3:
        return score

    sorted_days = sorted(day_frames.keys(), reverse=True)
    recent_days = sorted_days[:5]
    frames = [day_frames[day] for day in recent_days if not day_frames[day].empty]

    closes = [df["close"].iloc[-1] for df in frames]
    highs = [df["high"].max() for df in frames]
    lows = [df["low"].min() for df in frames]
    volumes = [df["volume"].sum() for df in frames]

    # ①トレンド：高値・安値が連続で切り下げ
    if len(highs) >= 3 and len(lows) >= 3:
        if highs[2] > highs[1] > highs[0] and lows[2] > lows[1] > lows[0]:
            score["trend"] = 2
        elif highs[1] > highs[0] or lows[1] > lows[0]:
            score["trend"] = 1

    # ②出来高：前々日急増 → 前日急減
    if len(volumes) >= 3 and volumes[1] < volumes[2] and volumes[1] < volumes[0]:
        score["volume"] = 2
    elif len(volumes) >= 2 and volumes[0] < volumes[1]:
        score["volume"] = 1

    # ③ブレイク位置：終値が前日安値を下回る
    if len(lows) >= 2 and len(closes) >= 1:
        diff = (closes[0] - lows[1]) / lows[1]
        if diff <= -0.005:
            score["break"] = 2
        elif diff < 0:
            score["break"] = 1

    # ④引け位置：終値が当日安値に近い
    today_low = lows[0]
    today_close = closes[0]
    if today_low:
        diff = (today_close - today_low) / today_low
        if diff <= 0.005:
            score["close_pos"] = 2
        elif diff <= 0.01:
            score["close_pos"] = 1

    # ⑤ボラティリティ
    today_high = highs[0]
    if today_close > 0 and (today_high - today_low) / today_close >= 0.03:
        score["volatility"] = 1

    # ⑥出来高水準
    if len(volumes) >= 5:
        avg_volume = sum(volumes[1:5]) / 4
        if volumes[0] >= avg_volume * 1.5:
            score["vol_level"] = 1

    return score


def create_score_table_long(data_dict):
    score_table = []

    for ticker, day_frames in data_dict.items():
        try:
            sorted_days = sorted(day_frames.keys())
            if len(sorted_days) < 2:
                continue
            # スコアは「前日」までで評価
            score_target_days = {day: day_frames[day] for day in sorted_days[:-1]}

            score = evaluate_stock_long(score_target_days)
            total = sum(score.values())

            # 前日のデータを表示用に使用
            prev_day = sorted_days[-2]
            prev_df = day_frames[prev_day]

            score_table.append(
                {
                    "ticker": ticker,
                    "終値": prev_df["close"].iloc[-1],
                    "直近高値": prev_df["high"].max(),
                    "直近安値": prev_df["low"].min(),
                    "トレンド": score["trend"],
                    "出来量変化": score["volume"],
                    "ブレイク": score["break"],
                    "引け位置": score["close_pos"],
                    "ボラ": score["volatility"],
                    "出来高水準": score["vol_level"],
                    "合計スコア": total,
                }
            )
        except Exception as e:
            print(f"{ticker}（ロング）処理中にエラー: {e}")

    df_score = pd.DataFrame(score_table)
    df_score = df_score.sort_values(by="合計スコア", ascending=False).reset_index(
        drop=True
    )
    return df_score


def create_score_table_short(data_dict):
    score_table = []

    for ticker, day_frames in data_dict.items():
        try:
            sorted_days = sorted(day_frames.keys())
            if len(sorted_days) < 2:
                continue
            score_target_days = {day: day_frames[day] for day in sorted_days[:-1]}

            score = evaluate_stock_short(score_target_days)
            total = sum(score.values())

            prev_day = sorted_days[-2]
            prev_df = day_frames[prev_day]

            score_table.append(
                {
                    "ticker": ticker,
                    "終値": prev_df["close"].iloc[-1],
                    "直近高値": prev_df["high"].max(),
                    "直近安値": prev_df["low"].min(),
                    "トレンド": score["trend"],
                    "出来量変化": score["volume"],
                    "ブレイク": score["break"],
                    "引け位置": score["close_pos"],
                    "ボラ": score["volatility"],
                    "出来高水準": score["vol_level"],
                    "合計スコア": total,
                }
            )
        except Exception as e:
            print(f"{ticker}（ショート）処理中にエラー: {e}")

    df_score = pd.DataFrame(score_table)
    df_score = df_score.sort_values(by="合計スコア", ascending=False).reset_index(
        drop=True
    )
    return df_score


def export_sheets(src_path, top_long, top_short, code_to_name):
    global EXCEL_PATH_L, EXCEL_PATH_S

    def process_copy(dst_path, code_list):
        import shutil
        import xlwings as xw

        shutil.copy(src_path, dst_path)
        app = xw.App(visible=False)
        app.display_alerts = False
        wb = app.books.open(dst_path)

        # コードからシート名（銘柄名）に変換
        sheet_name_list = [code_to_name.get(code, "") for code in code_list]

        for sheet in tqdm(wb.sheets, desc="シート削除中"):
            if sheet.name not in sheet_name_list:
                try:
                    sheet.delete()
                except Exception as e:
                    print(f"⚠️ シート {sheet.name} の削除に失敗: {e}")
            else:
                formula = sheet.range("A1").formula
                if (
                    isinstance(formula, str)
                    and formula.startswith("=RssChart")
                    and f", {RSS_PARAM_TO_REPLACE}" in formula
                ):
                    sheet.range("A1").formula = formula.replace(
                        f", {RSS_PARAM_TO_REPLACE}", f", {RSS_PARAM_NEW}"
                    )

        wb.save()
        wb.close()
        app.quit()

    if top_long is not None and not top_long.empty:
        print(f"📊 買いスコア上位:\n{top_long}")
        process_copy(EXCEL_PATH_L, top_long["ticker"].tolist())

    if top_short is not None and not top_short.empty:
        print(f"📉 売りスコア上位:\n{top_short}")
        process_copy(EXCEL_PATH_S, top_short["ticker"].tolist())


def export_top_sheets():
    src_path = "デイトレ株価データ.xlsx"

    # スコア表読み込み
    long_df = pd.read_csv("score_table_long.csv", encoding="shift_jis")
    short_df = pd.read_csv("score_table_short.csv", encoding="shift_jis")

    # スコア7点以上のみ抽出
    top_long = long_df[long_df["合計スコア"] >= SCORE_THRESHOLD_L]["ticker"].tolist()
    top_short = short_df[short_df["合計スコア"] >= SCORE_THRESHOLD_S]["ticker"].tolist()

    # Excel起動
    app = xw.App(visible=False)
    wb_src = app.books.open(src_path)

    # ✅ 買い銘柄ファイルの作成
    if top_long:
        wb_long = app.books.add()
        for sheet_name in top_long:
            if sheet_name in [s.name for s in wb_src.sheets]:
                wb_src.sheets[sheet_name].copy(after=wb_long.sheets[-1])
            else:
                print(f"⚠️ 買いシート {sheet_name} が見つかりません")
        if len(wb_long.sheets) > 1 and wb_long.sheets[0].name == "Sheet1":
            wb_long.sheets[0].delete()
        for sheet in wb_long.sheets:
            formula = sheet.range("A1").formula
            if (
                isinstance(formula, str)
                and formula.startswith("=RssChart")
                and f", {RSS_PARAM_TO_REPLACE}" in formula
            ):
                sheet.range("A1").formula = formula.replace(
                    f", {RSS_PARAM_TO_REPLACE}", f", {RSS_PARAM_NEW}"
                )
        wb_long.save("買い銘柄寄り後情報.xlsx")
        wb_long.close()

    # ✅ 売り銘柄ファイルの作成
    if top_short:
        wb_short = app.books.add()
        for sheet_name in top_short:
            if sheet_name in [s.name for s in wb_src.sheets]:
                wb_src.sheets[sheet_name].copy(after=wb_short.sheets[-1])
            else:
                print(f"⚠️ 売りシート {sheet_name} が見つかりません")
        if len(wb_short.sheets) > 1 and wb_short.sheets[0].name == "Sheet1":
            wb_short.sheets[0].delete()
        for sheet in wb_short.sheets:
            formula = sheet.range("A1").formula
            if (
                isinstance(formula, str)
                and formula.startswith("=RssChart")
                and f", {RSS_PARAM_TO_REPLACE}" in formula
            ):
                sheet.range("A1").formula = formula.replace(
                    f", {RSS_PARAM_TO_REPLACE}", f", {RSS_PARAM_NEW}"
                )
        wb_short.save("売り銘柄寄り後情報.xlsx")
        wb_short.close()

    wb_src.close()
    app.quit()


# 実行
if __name__ == "__main__":
    print("デイトレ株価データからスコア表を作成します")
    excel_path = "デイトレ株価データ.xlsx"
    data_dict = load_data(excel_path)

    # 買い候補（ロング）
    result_long = create_score_table_long(data_dict)
    result_long.to_csv("score_table_long.csv", index=False, encoding="shift_jis")

    # 売り候補（ショート）
    result_short = create_score_table_short(data_dict)
    result_short.to_csv("score_table_short.csv", index=False, encoding="shift_jis")

    export_top_sheets()
    print("スコア表作成完了")
