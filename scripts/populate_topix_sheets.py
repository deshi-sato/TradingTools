\"\"\"populate_topix_sheets.py : Auto-generated placeholder

- file: scripts/populate_topix_sheets.py
- updated: 2025-09-08

TODO: このモジュールの概要をここに書いてください。
\"\"\"
import sys
from pathlib import Path
from typing import List
import shutil

from openpyxl import load_workbook


def read_codes(path: Path) -> List[str]:
    if not path.exists():
        raise FileNotFoundError(f"Codes file not found: {path}")
    codes: List[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        code = line.strip()
        if code:
            codes.append(code)
    if not codes:
        raise ValueError("No codes found in the codes file.")
    return codes


def ensure_100_sheets_and_fill_q1(xlsm_path: Path, codes_path: Path) -> None:
    if not xlsm_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsm_path}")

    # Backup original workbook once before modifying
    backup_path = xlsm_path.with_suffix(xlsm_path.suffix + ".bak")
    if not backup_path.exists():
        shutil.copyfile(xlsm_path, backup_path)

    codes = read_codes(codes_path)
    if len(codes) < 100:
        raise ValueError(f"Expected at least 100 codes, found {len(codes)} in {codes_path}")

    wb = load_workbook(filename=str(xlsm_path), keep_vba=True)

    if len(wb.worksheets) == 0:
        raise ValueError("Workbook has no worksheets.")

    base_sheet = wb.worksheets[0]

    # Create copies so that total sheets == 100
    total_needed = 100
    current = len(wb.worksheets)
    if current < total_needed:
        base_title = base_sheet.title
        for i in range(current, total_needed):
            new_ws = wb.copy_worksheet(base_sheet)
            # Assign a deterministic unique title for clarity
            new_ws.title = f"{base_title}_{i+1}"

    # Fill Q1 with the first 100 codes (one per sheet)
    for idx in range(100):
        ws = wb.worksheets[idx]
        ws["Q1"] = codes[idx]

    wb.save(str(xlsm_path))


def main(argv: List[str]) -> int:
    # Defaults based on the user request
    xlsm_path = Path("rss_snapshot.xlsm")
    codes_path = Path("data/topix100_codes.txt")
    try:
        ensure_100_sheets_and_fill_q1(xlsm_path, codes_path)
    except Exception as e:
        print(f"Error: {e}")
        return 1
    print("Done: Created 100 sheets and populated Q1 from codes.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

