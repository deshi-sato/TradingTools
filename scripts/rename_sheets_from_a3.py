\"\"\"rename_sheets_from_a3.py : Auto-generated placeholder

- file: scripts/rename_sheets_from_a3.py
- updated: 2025-09-08

TODO: このモジュールの概要をここに書いてください。
\"\"\"
import sys
from pathlib import Path
import re
import shutil
from typing import Set

from openpyxl import load_workbook


INVALID_CHARS_RE = re.compile(r"[:\\/?*\[\]]")
MAX_LEN = 31


def sanitize_name(name: str) -> str:
    name = name.strip()
    name = INVALID_CHARS_RE.sub("_", name)
    # Excel sheet names cannot begin or end with a single quote
    name = name.strip("'")
    # Collapse whitespace
    name = re.sub(r"\s+", " ", name)
    return name[:MAX_LEN]


def unique_name(base: str, used: Set[str]) -> str:
    if base and base not in used:
        return base
    if not base:
        base = "Sheet"
    # Try with numeric suffixes
    n = 2
    while True:
        suffix = f"_{n}"
        max_base_len = MAX_LEN - len(suffix)
        candidate = base[:max_base_len] + suffix
        if candidate not in used:
            return candidate
        n += 1


def rename_sheets_from_a3(xlsm_path: Path) -> None:
    if not xlsm_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsm_path}")

    # Backup once (non-destructive if already exists)
    backup_path = xlsm_path.with_suffix(xlsm_path.suffix + ".bak.rename")
    if not backup_path.exists():
        shutil.copyfile(xlsm_path, backup_path)

    # data_only=True to read cached formula results if A3 contains formulas
    wb = load_workbook(filename=str(xlsm_path), keep_vba=True, data_only=True)

    used = set()
    rename_map = {}
    skipped = []
    for ws in wb.worksheets:
        raw = ws["A3"].value
        proposed = (str(raw) if raw is not None else "").strip()
        if not proposed:
            skipped.append(ws.title)
            continue
        proposed = sanitize_name(proposed)
        new_title = unique_name(proposed, used)
        used.add(new_title)
        rename_map[ws.title] = new_title

    # First, rename all target sheets to temporary unique names to avoid collisions
    temp_suffix = "__TMP__"
    temp_names = {}
    for ws in wb.worksheets:
        old = ws.title
        if old in rename_map:
            tmp = unique_name(old[: MAX_LEN - len(temp_suffix)] + temp_suffix, set(w.title for w in wb.worksheets))
            ws.title = tmp
            temp_names[tmp] = old

    # Second, apply the final names
    for ws in wb.worksheets:
        # Map back from temp to original old name
        orig_old = temp_names.get(ws.title)
        if orig_old and orig_old in rename_map:
            ws.title = rename_map[orig_old]

    wb.save(str(xlsm_path))
    print("Renamed sheets:")
    for old, new in rename_map.items():
        print(f"  {old} -> {new}")
    if skipped:
        print("Skipped (A3 empty):")
        for name in skipped:
            print(f"  {name}")


def main() -> int:
    xlsm_path = Path("rss_snapshot.xlsm")
    try:
        rename_sheets_from_a3(xlsm_path)
    except Exception as e:
        print(f"Error: {e}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
