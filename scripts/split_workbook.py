\"\"\"split_workbook.py : Auto-generated placeholder

- file: scripts/split_workbook.py
- updated: 2025-09-08

TODO: このモジュールの概要をここに書いてください。
\"\"\"
from pathlib import Path
from typing import Iterable, Set
from openpyxl import load_workbook


def remove_all_except_indices(wb, keep_indices: Iterable[int]) -> None:
    keep: Set[int] = set(keep_indices)
    # Build list first to avoid mutating while iterating
    to_remove = [ws for idx, ws in enumerate(wb.worksheets) if idx not in keep]
    for ws in to_remove:
        wb.remove(ws)
    # Make the first remaining sheet active (if any)
    if wb.worksheets:
        wb.active = 0


def split_workbook(src: Path, out1: Path, out2: Path) -> None:
    if not src.exists():
        raise FileNotFoundError(f"Workbook not found: {src}")

    # First half: keep first 50
    wb1 = load_workbook(filename=str(src), keep_vba=True)
    n = len(wb1.worksheets)
    if n < 50:
        raise ValueError(f"Expected at least 50 sheets, found {n}")
    print(f"Loaded workbook: {n} sheets")
    remove_all_except_indices(wb1, range(0, 50))
    print("Saving first part...")
    wb1.save(str(out1))
    print(f"Saved: {out1}")

    # Second half: keep remaining (50..end)
    wb2 = load_workbook(filename=str(src), keep_vba=True)
    n2 = len(wb2.worksheets)
    remove_all_except_indices(wb2, range(50, n2))
    print("Saving second part...")
    wb2.save(str(out2))
    print(f"Saved: {out2}")


def main() -> int:
    src = Path("rss_snapshot.xlsm")
    out1 = Path("rss_snapshot_part1.xlsm")
    out2 = Path("rss_snapshot_part2.xlsm")
    try:
        split_workbook(src, out1, out2)
    except Exception as e:
        print(f"Error: {e}")
        return 1
    print(f"Wrote {out1} and {out2}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
