\"\"\"update_indices_fixed.py : Auto-generated placeholder

- file: scripts/update_indices_fixed.py
- updated: 2025-09-08

TODO: このモジュールの概要をここに書いてください。
\"\"\"
# -*- coding: utf-8 -*-
r"""
固定セルの指標データを Excel(.xlsm) から読み、SQLite に書き込む。
- 既定セル配置はハードコード／--config で上書き可／名前付き範囲も可
- --interval で常時取り込み（Ctrl+Cで停止）
- --mode replace|append で保存モード選択
    replace: 最新だけ残す（毎回クリア）
    append : 既存保持＋差分だけ追記（重複なし）
"""

import argparse, sqlite3, sys, time, os, random, re
from scripts.common_config import load_json_utf8
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta
from typing import Tuple, Optional, Dict, Any, List

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

# ================= 既定のセル配置 =================
DEFAULT_MAPPING = {
    "sheet": 1,
    "snapshots": [
        {"code": "DJIA",   "name": "NYダウ",           "date": "B3",  "time": "C3",  "last": "D3",  "pct": "E3"},
        {"code": "NASDAQ", "name": "NASDAQ総合指数",   "date": "B4",  "time": "C4",  "last": "D4",  "pct": "E4"},
        {"code": "SP500",  "name": "S&P500指数",       "date": "B5",  "time": "C5",  "last": "D5",  "pct": "E5"},
        {"code": "VIX",    "name": "VIX指数",          "date": "B6",  "time": "C6",  "last": "D6",  "pct": "E6"},
    ],
    "ohlcv": [
        {"code":"USDJPY","name":"ドル/円(Bid)","rows":100,
         "date":"H3","time":"I3","open":"J3","high":"K3","low":"L3","close":"M3","volume":"N3"},
        {"code":"N225","name":"日経225","rows":100,
         "date":"Q3","time":"R3","open":"S3","high":"T3","low":"U3","close":"V3","volume":"W3"},
        {"code":"N225_FUT","name":"225先物大阪(期近)","rows":100,
         "date":"Z3","time":"AA3","open":"AB3","high":"AC3","low":"AD3","close":"AE3","volume":"AF3"},
        {"code":"TOPIX","name":"TOPIX","rows":100,
         "date":"AI3","time":"AJ3","open":"AK3","high":"AL3","low":"AM3","close":"AN3","volume":"AO3"},
        {"code":"NikkeiVI","name":"日経平均VI指数","rows":100,
         "date":"AR3","time":"AS3","open":"AT3","high":"AU3","low":"AV3","close":"AW3","volume":"AX3"}
    ]
}

# ================ DB スキーマ =================
DDL = """
CREATE TABLE IF NOT EXISTS market_snapshots(
  code TEXT NOT NULL,
  datetime TEXT NOT NULL,     -- 'YYYY-MM-DD HH:MM:SS' JST
  last REAL,
  pct_change REAL,
  source TEXT DEFAULT 'xlsm',
  PRIMARY KEY(code, datetime)
);
CREATE TABLE IF NOT EXISTS market_ohlcv(
  code TEXT NOT NULL,
  datetime TEXT NOT NULL,     -- 'YYYY-MM-DD HH:MM:SS' JST
  open REAL, high REAL, low REAL, close REAL, volume REAL,
  source TEXT DEFAULT 'xlsm',
  PRIMARY KEY(code, datetime)
);
CREATE INDEX IF NOT EXISTS idx_ms_code_dt ON market_snapshots(code, datetime);
CREATE INDEX IF NOT EXISTS idx_mo_code_dt ON market_ohlcv(code, datetime);
"""

# ================ Excelユーティリティ =================
EXCEL_EPOCH = datetime(1899, 12, 30)
_HYPHEN_RE = re.compile(r"^-{3,}$")

def coords_from_mapping(ws, it):
    """it = {'date':'H3', 'time':'I3', 'open':'J3', ...} を行頭/各列indexに分解"""
    r_date, c_date = coordinate_to_tuple(it["date"])
    r_time, c_time = coordinate_to_tuple(it["time"])
    # 先頭行（スタート行）は date/time のどちらでもOKだが低い方に合わせる
    r0 = min(r_date, r_time)
    cols = (
        c_date, c_time,
        coordinate_to_tuple(it["open"])[1],
        coordinate_to_tuple(it["high"])[1],
        coordinate_to_tuple(it["low"])[1],
        coordinate_to_tuple(it["close"])[1],
        coordinate_to_tuple(it["volume"])[1],
    )
    return r0, cols  # 行開始位置, (date,time,open,high,low,close,vol) 各列番号

def iter_rows_until_hyphen(ws, start_row, cols, max_rows=2000):
    """行ベースで '-----' が来るまで (date,time,o,h,l,c,v) のタプルをyield"""
    # openpyxlは列範囲が離れていてもOK。min/maxで矩形読みして抽出するのが最速。
    min_col, max_col = min(cols), max(cols)
    # values_only=True でセルオブジェクト生成を避ける
    for i, row in enumerate(ws.iter_rows(min_row=start_row, min_col=min_col, max_col=max_col,
                                         values_only=True)):
        if i >= max_rows:
            break
        # 欲しい列だけを取り出す
        date = row[cols[0]-min_col]
        time_v = row[cols[1]-min_col]
        if is_hyphen_sentinel(date) or is_hyphen_sentinel(time_v):
            break
        o = row[cols[2]-min_col]
        h = row[cols[3]-min_col]
        l = row[cols[4]-min_col]
        c = row[cols[5]-min_col]
        v = row[cols[6]-min_col]
        yield (date, time_v, o, h, l, c, v)

def is_hyphen_sentinel(x) -> bool:
    """'-----' などの終端行か？"""
    if x is None:
        return False
    return bool(_HYPHEN_RE.fullmatch(str(x).strip()))

def read_col_until_sentinel(ws: Worksheet, start_addr_or_name: str, max_rows: int = 2000):
    """
    start セルから下方向に、'-----' が出るまで読み続けてリストで返す。
    max_rows は安全上限（デフォルト 2000）。
    """
    # 名前付き範囲→アドレス化
    r0 = c0 = None
    try:
        dn = ws.parent.defined_names.get(start_addr_or_name)
        if dn is not None:
            for title, ref in dn.destinations:
                if title == ws.title:
                    ref = ref.split(":")[0]
                    r0, c0 = coordinate(ref)
    except Exception:
        pass
    if r0 is None:
        r0, c0 = coordinate(start_addr_or_name)

    out = []
    for i in range(max_rows):
        v = ws.cell(row=r0 + i, column=c0).value
        if is_hyphen_sentinel(v):
            break
        out.append(v)
    return out

def parse_excel_date(val) -> Optional[datetime]:
    if val is None or val == "": return None
    if isinstance(val, datetime): return val
    try:
        f = float(val); return EXCEL_EPOCH + timedelta(days=f)
    except Exception:
        pass
    for fmt in ("%Y-%m-%d","%Y/%m/%d","%y/%m/%d","%Y%m%d"):
        try: return datetime.strptime(str(val).strip(), fmt)
        except Exception: continue
    return None

def parse_excel_time(val) -> Optional[datetime]:
    if val is None or val == "": return None
    if isinstance(val, datetime): return val
    try:
        f = float(val); return EXCEL_EPOCH + timedelta(days=f)
    except Exception:
        pass
    for fmt in ("%H:%M:%S","%H:%M"):
        try: return datetime.strptime(str(val).strip(), fmt)
        except Exception: continue
    return None

def parse_percent(val) -> Optional[float]:
    if val is None or val == "": return None
    s = str(val).strip()
    if s.endswith("%"):
        try: return float(s[:-1]) / 100.0
        except Exception: return None
    try:
        f = float(s); return f if abs(f) <= 1.0 else f/100.0
    except Exception: return None

def coordinate(addr: str) -> Tuple[int,int]:
    r, c = coordinate_to_tuple(addr); return int(r), int(c)

def resolve_sheet(wb, spec):
    if isinstance(spec, int): return wb.worksheets[spec-1]
    if isinstance(spec, str) and spec.isdigit(): return wb.worksheets[int(spec)-1]
    return wb[spec]

def get_named_or_cell(ws: Worksheet, name_or_addr: str):
    try:
        dn = ws.parent.defined_names.get(name_or_addr)
        if dn is not None:
            for title, ref in dn.destinations:
                if title == ws.title:
                    ref = ref.split(":")[0]
                    r, c = coordinate(ref)
                    return ws.cell(row=r, column=c).value
    except Exception:
        pass
    r, c = coordinate(name_or_addr)
    return ws.cell(row=r, column=c).value

def read_down(ws: Worksheet, start_addr_or_name: str, n: int):
    r0 = c0 = None
    try:
        dn = ws.parent.defined_names.get(start_addr_or_name)
        if dn is not None:
            for title, ref in dn.destinations:
                if title == ws.title:
                    ref = ref.split(":")[0]
                    r0, c0 = coordinate(ref)
    except Exception:
        pass
    if r0 is None:
        r0, c0 = coordinate(start_addr_or_name)
    return [ws.cell(row=r0+i, column=c0).value for i in range(n)]

# ================ ファイルをロックしない安全読み込み =================
def read_file_safely(path: Path, stable_wait=0.2, retries=5, backoff=0.3) -> bytes:
    last_err = None
    for _ in range(retries):
        try:
            s1 = os.stat(path); time.sleep(stable_wait); s2 = os.stat(path)
            if s1.st_mtime != s2.st_mtime or s1.st_size != s2.st_size:
                time.sleep(backoff); continue
            with open(path, "rb") as f: return f.read()
        except Exception as e:
            last_err = e; time.sleep(backoff)
    if last_err: raise last_err
    raise RuntimeError(f"failed to read stable file: {path}")

def open_workbook_no_lock(path: Path):
    data = read_file_safely(path)
    return load_workbook(filename=BytesIO(data), read_only=True, data_only=True)

# ================ 共通：ヘルパ =================
def is_hyphen_sentinel(x) -> bool:
    """ '-----' などの終端行か？ """
    if x is None: return False
    s = str(x).strip()
    return bool(_HYPHEN_RE.fullmatch(s))

def dt_to_str(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def get_last_dt(con: sqlite3.Connection, code: str) -> Optional[datetime]:
    cur = con.execute("SELECT MAX(datetime) FROM market_ohlcv WHERE code=?", (code,))
    row = cur.fetchone()
    if row and row[0]:
        try: return datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
        except Exception: return None
    return None

# ================ 取り込みルーチン =================
def import_snapshots(ws: Worksheet, con: sqlite3.Connection, items, mode: str):
    cur = con.cursor(); added = 0; per_code = []
    for it in items:
        code = it["code"]
        dval = get_named_or_cell(ws, it["date"])
        tval = get_named_or_cell(ws, it["time"])
        lval = get_named_or_cell(ws, it["last"])
        pval = get_named_or_cell(ws, it["pct"])

        d = parse_excel_date(dval); t = parse_excel_time(tval)
        if d is None and t is None:
            per_code.append((code, 0)); continue
        if d and t: dt = d.replace(hour=t.hour, minute=t.minute, second=getattr(t, "second", 0))
        elif d:     dt = d
        else:
            now = datetime.now(); dt = now.replace(hour=t.hour, minute=t.minute, second=getattr(t, "second", 0), microsecond=0)

        last = float(lval) if lval not in (None,"") else None
        pct  = parse_percent(pval)

        cur.execute("""
        INSERT OR REPLACE INTO market_snapshots(code, datetime, last, pct_change, source)
        VALUES(?,?,?,?, 'xlsm')
        """, (code, dt_to_str(dt), last, pct))
        added += 1; per_code.append((code, 1))
    con.commit()
    return added, per_code

def import_ohlcv(ws: Worksheet, con: sqlite3.Connection, items, mode: str):
    cur = con.cursor(); total = 0; per_code = []
    for it in items:
        code = it["code"]
        last_dt = get_last_dt(con, code) if mode == "append" else None

        start_row, cols = coords_from_mapping(ws, it)

        added = 0
        buf = []
        for date_v, time_v, o_v, h_v, l_v, c_v, vol_v in iter_rows_until_hyphen(ws, start_row, cols, max_rows=5000):
            d = parse_excel_date(date_v)
            t = parse_excel_time(time_v)
            if d is None and t is None:
                continue
            if d and t:
                dt = d.replace(hour=t.hour, minute=t.minute, second=getattr(t, "second", 0))
            elif d:
                dt = d
            else:
                now = datetime.now()
                dt = now.replace(hour=t.hour, minute=t.minute, second=getattr(t, "second", 0), microsecond=0)

            if last_dt is not None and not (dt > last_dt):
                continue

            try:
                o = float(o_v) if o_v not in (None,"") else None
                h = float(h_v) if h_v not in (None,"") else None
                l = float(l_v) if l_v not in (None,"") else None
                c = float(c_v) if c_v not in (None,"") else None
                v = float(vol_v) if vol_v not in (None,"") else None
            except Exception:
                o=h=l=c=v=None

            buf.append((dt.strftime("%Y-%m-%d %H:%M:%S"), o, h, l, c, v))

        if buf:
            cur.executemany(
                """
                INSERT OR REPLACE INTO market_ohlcv
                  (code, datetime, open, high, low, close, volume, source)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'xlsm')
                """,
                [(code, *r) for r in buf]
            )
            con.commit()
            added = len(buf)

        total += added
        per_code.append((code, added))
    return total, per_code

# ================ 実行本体 =================
def run_once(excel_path: Path, db_path: Path, mapping: Dict[str,Any], sheet_override=None, mode="replace"):
    wb = open_workbook_no_lock(excel_path)  # ファイルは即解放
    try:
        ws = resolve_sheet(wb, sheet_override if sheet_override is not None else mapping.get("sheet", 1))
        con = sqlite3.connect(db_path)
        con.executescript(DDL)
        con.execute("PRAGMA journal_mode=WAL;")
        con.execute("PRAGMA synchronous=NORMAL;")

        with con:
            if mode == "replace":
                # 最新だけ残す：毎回クリア
                con.execute("DELETE FROM market_snapshots;")
                con.execute("DELETE FROM market_ohlcv;")

            n1, snap_detail = import_snapshots(ws, con, mapping.get("snapshots", []), mode)
            n2, ohlc_detail = import_ohlcv(ws, con, mapping.get("ohlcv", []), mode)

        con.close()
        return n1, snap_detail, n2, ohlc_detail
    finally:
        try: wb.close()
        except Exception: pass

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="指標データ.xlsm のパス")
    ap.add_argument("--db", default=r".\data\rss_index.db", help="出力DB（既定: .\\data\\rss_index.db）")
    ap.add_argument("--sheet", default=None, help="シート（名 or 1始まり番号）")
    ap.add_argument("--config", default=None, help="セル配置JSON。渡せば既定を置き換え")
    ap.add_argument("--interval", type=int, default=None, help="秒数。指定時はこの間隔で継続実行")
    ap.add_argument("--once", action="store_true", help="1回だけ実行して終了")
    ap.add_argument("--mode", choices=["replace","append"], default="replace",
                    help="replace:毎回クリアして最新のみ保持 / append:既存保持して新行だけ追記（重複なし）")
    args = ap.parse_args()

    excel_path = Path(args.excel); db_path = Path(args.db)
    if not excel_path.exists(): raise FileNotFoundError(f"Excel が見つかりません: {excel_path}")

    mapping = DEFAULT_MAPPING
    if args.config:
        mapping = load_json_utf8(args.config)

    # 単発
    if args.once or not args.interval:
        n1, sd, n2, od = run_once(excel_path, db_path, mapping, args.sheet, args.mode)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{now}] snapshots={n1} rows  ohlcv={n2} rows  → {db_path.resolve()}")
        if sd: print("  snapshots:", ", ".join([f"{c}:{k}" for c,k in sd]))
        if od: print("  ohlcv    :", ", ".join([f"{c}:{k}" for c,k in od]))
        return

    # ループ実行
    interval = max(1, int(args.interval))
    print(f"Start loop: excel='{excel_path}', db='{db_path}', mode={args.mode}, every {interval}s  (Ctrl+C to stop)")
    try:
        while True:
            loop_start = time.time()
            try:
                n1, sd, n2, od = run_once(excel_path, db_path, mapping, args.sheet, args.mode)
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                snap_str = " ".join([f"{c}:{k}" for c,k in sd]) if sd else "-"
                ohlc_str = " ".join([f"{c}:{k}" for c,k in od]) if od else "-"
                print(f"[{now}] SNAP={n1} ({snap_str}) | OHLCV={n2} ({ohlc_str})")
            except Exception as e:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                print(f"[{now}] ERROR: {e}", file=sys.stderr)

            elapsed = time.time() - loop_start
            to_sleep = max(0, interval - elapsed)
            time.sleep(to_sleep + random.uniform(0, 0.3))
    except KeyboardInterrupt:
        print("\nStopped by user.")

if __name__ == "__main__":
    main()
