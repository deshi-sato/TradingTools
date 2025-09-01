# -*- coding: utf-8 -*-
"""
watchlist_YYYY-MM-DD.csv の最新ファイルから、
シート 'index' の A1..A15 に BUY、A16..A30 に SELL の銘柄コードを書き込む。

- CSV 想定列: side, tickerｄ (←全角d), または tickerd / ticker / Ticker / TICKER
- 値は '1925.T' のような形式 → '1925' に整形
- BUY/SELL は CSV の順番を尊重、重複は削除しない
- .xlsm は keep_vba=True で保存
"""

from __future__ import annotations
import argparse
import logging
import re
from pathlib import Path
from typing import List, Tuple

import pandas as pd
from openpyxl import load_workbook

# 既定値
DEFAULT_DATA_DIR = Path("./data")
DEFAULT_EXCEL = DEFAULT_DATA_DIR / "株価データ.xlsm"
DEFAULT_SHEET = "index"
DEFAULT_MAX_BUY = 15
DEFAULT_MAX_SELL = 15

WATCH_RE = re.compile(r"watchlist_(\d{4}-\d{2}-\d{2})\.csv$", re.IGNORECASE)


def setup_logger(log_path: str | None, level: str = "INFO") -> logging.Logger:
    logger = logging.getLogger("rss_index")
    logger.handlers.clear()
    logger.setLevel(getattr(logging, level.upper(), logging.INFO))
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    ch = logging.StreamHandler()
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    if log_path:
        fh = logging.FileHandler(log_path, encoding="utf-8")
        fh.setFormatter(fmt)
        logger.addHandler(fh)
    return logger


def pick_latest_watchlist(data_dir: Path) -> Path:
    """ファイル名の日付を優先、なければ更新時刻で最新を選ぶ。"""
    cands = list(data_dir.glob("watchlist_*.csv"))
    if not cands:
        raise FileNotFoundError(f"{data_dir} に watchlist_*.csv が見つかりません。")
    dated: list[tuple[str, Path]] = []
    others: list[Path] = []
    for p in cands:
        m = WATCH_RE.match(p.name)
        if m:
            dated.append((m.group(1), p))  # 日付文字列の比較でOK（YYYY-MM-DD）
        else:
            others.append(p)
    if dated:
        dated.sort(key=lambda x: x[0])
        return dated[-1][1]
    return max(cands, key=lambda p: p.stat().st_mtime)


def read_codes_by_side(
    path: Path, max_buy: int, max_sell: int, logger: logging.Logger
) -> Tuple[List[str], List[str]]:
    # 文字コード差異に強く
    df = pd.read_csv(path, encoding="utf-8-sig")
    cols_lower = {c.lower(): c for c in df.columns}

    # side 列
    if "side" not in cols_lower:
        raise KeyError(f"{path.name} に side 列がありません。実列: {list(df.columns)}")
    side_col = cols_lower["side"]

    # ティッカー列（全角d優先）
    tick_candidates = ["tickerｄ", "tickerd", "ticker", "Ticker", "TICKER"]
    tick_col = None
    for c in tick_candidates:
        if c in df.columns:
            tick_col = c
            break
    if tick_col is None:
        # 小文字化でも探索
        for c in df.columns:
            if c.lower().startswith("ticker"):
                tick_col = c
                break
    if tick_col is None:
        raise KeyError(
            f"{path.name} にティッカー列が見つかりません。候補: {tick_candidates} / 実列: {list(df.columns)}"
        )

    def to_code(x) -> str:
        s = str(x).strip()
        return s.split(".", 1)[0] if s else ""

    buy_all = [
        to_code(x)
        for x in df.loc[
            df[side_col].astype(str).str.upper() == "BUY", tick_col
        ].tolist()
    ]
    sell_all = [
        to_code(x)
        for x in df.loc[
            df[side_col].astype(str).str.upper() == "SELL", tick_col
        ].tolist()
    ]

    buys = buy_all[:max_buy]
    sells = sell_all[:max_sell]

    logger.info(
        "watchlist=%s  BUY(raw)=%d -> %d  SELL(raw)=%d -> %d",
        path.name,
        len(buy_all),
        len(buys),
        len(sell_all),
        len(sells),
    )
    if not buys and not sells:
        logger.warning(
            "BUY/SELL ともに空です。csv の side 列や列名を確認してください。"
        )
    return buys, sells


def write_to_xlsm(
    excel_path: Path,
    sheet: str,
    buys: List[str],
    sells: List[str],
    max_buy: int,
    max_sell: int,
    logger: logging.Logger,
) -> None:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel が見つかりません: {excel_path}")
    wb = load_workbook(
        filename=str(excel_path), keep_vba=True, read_only=False, data_only=False
    )
    try:
        if sheet not in wb.sheetnames:
            raise KeyError(f"シート '{sheet}' がありません。存在: {wb.sheetnames}")
        ws = wb[sheet]

        # A1..A{max_buy} に BUY、A{max_buy+1}..A{max_buy+max_sell} に SELL
        for i in range(max_buy):
            ws.cell(row=i + 1, column=1, value=(buys[i] if i < len(buys) else None))
        for j in range(max_sell):
            ws.cell(
                row=max_buy + j + 1,
                column=1,
                value=(sells[j] if j < len(sells) else None),
            )

        wb.save(str(excel_path))
        logger.info(
            "wrote %s!A1:A%d (BUY:%d, SELL:%d)",
            sheet,
            max_buy + max_sell,
            len(buys),
            len(sells),
        )
    finally:
        try:
            wb.close()
        except Exception:
            pass


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--data-dir",
        default=str(DEFAULT_DATA_DIR),
        help="watchlist_*.csv のディレクトリ",
    )
    ap.add_argument(
        "--excel", default=str(DEFAULT_EXCEL), help="rss_snapshot.xlsm のパス"
    )
    ap.add_argument(
        "--sheet", default=DEFAULT_SHEET, help="書き込み先シート名（既定: index）"
    )
    ap.add_argument("--max-buy", type=int, default=DEFAULT_MAX_BUY)
    ap.add_argument("--max-sell", type=int, default=DEFAULT_MAX_SELL)
    ap.add_argument("--log", help="ログファイル（未指定ならコンソールのみ）")
    ap.add_argument("--log-level", default="INFO")
    args = ap.parse_args()

    logger = setup_logger(args.log, args.log_level)

    data_dir = Path(args.data_dir)
    excel_path = Path(args.excel)

    wl = pick_latest_watchlist(data_dir)
    logger.info("picked watchlist: %s", wl.name)

    buys, sells = read_codes_by_side(wl, args.max_buy, args.max_sell, logger)
    write_to_xlsm(
        excel_path, args.sheet, buys, sells, args.max_buy, args.max_sell, logger
    )
    logger.info("DONE.")


if __name__ == "__main__":
    main()
