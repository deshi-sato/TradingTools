from openpyxl import load_workbook
from pathlib import Path
import sys

p = Path('stock_data.xlsm')
try:
    wb = load_workbook(filename=str(p), keep_vba=True, read_only=False, data_only=False)
    print('OK', p, wb.sheetnames)
    wb.close()
except Exception as e:
    print('ERR', e)
    sys.exit(2)
